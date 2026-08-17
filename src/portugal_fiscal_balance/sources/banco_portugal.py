"""Banco de Portugal / INE long-series parser for historical public accounts."""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Final

import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

ESCUDOS_PER_EURO: Final[float] = 200.482

SECTOR_SHEETS: Final[dict[str, str]] = {
    "RDAP": "general_government",
    "RDAC": "central_government",
    "RDARL": "regional_local_government",
    "RDSS": "social_security_funds",
}

ADDITIONAL_BALANCE_SHEETS: Final[dict[str, str]] = {
    "RDE": "state_balance_m_eur",
    "RDAL": "local_government_balance_m_eur",
    "RDAR": "regional_government_balance_m_eur",
}


@dataclass(frozen=True)
class HistoricalExtraction:
    """Historical balances, account components, transfers and macro controls."""

    balances: pd.DataFrame
    accounts: pd.DataFrame
    transfers: pd.DataFrame
    macro: pd.DataFrame


def _normalise(value: object) -> str:
    return " ".join(str(value or "").lower().replace("|", " ").split())


def _year_columns(ws: Worksheet, row: int) -> dict[int, int]:
    result: dict[int, int] = {}
    for col in range(1, ws.max_column + 1):
        value = ws.cell(row, col).value
        if isinstance(value, (int, float)) and int(value) == value:
            year = int(value)
            if 1900 <= year <= 2100:
                result[year] = col
    if not result:
        raise ValueError(f"No year columns found in {ws.title!r}")
    return result


def _rows_containing(ws: Worksheet, needle: str) -> list[int]:
    target = _normalise(needle)
    rows: list[int] = []
    for row in range(1, ws.max_row + 1):
        values = [_normalise(ws.cell(row, col).value) for col in range(1, min(ws.max_column, 4) + 1)]
        if any(target in value for value in values):
            rows.append(row)
    return rows


def _row(ws: Worksheet, needle: str, *, occurrence: int = 0) -> int:
    rows = _rows_containing(ws, needle)
    if occurrence >= len(rows):
        raise ValueError(
            f"Could not find occurrence {occurrence} of {needle!r} in sheet {ws.title!r}"
        )
    return rows[occurrence]


def _to_m_eur(value: object) -> float | None:
    if value is None:
        return None
    if not isinstance(value, (int, float)):
        return None
    return float(value) / ESCUDOS_PER_EURO


def _extract_gdp(workbook_path: Path) -> pd.DataFrame:
    workbook = load_workbook(workbook_path, data_only=True, read_only=False)
    ws = workbook["PIBpm"]
    years = _year_columns(ws, 5)
    gdp_row = _row(ws, "PIBpm")
    records: list[dict[str, float | int]] = []
    for year, col in sorted(years.items()):
        if 1977 <= year <= 1995:
            value = _to_m_eur(ws.cell(gdp_row, col).value)
            if value is not None:
                records.append({"year": year, "nominal_gdp_m_eur": value})
    return pd.DataFrame.from_records(records)


def _extract_macro(workbook_path: Path) -> pd.DataFrame:
    """Extract historical labour-market controls published in the same workbook."""
    workbook = load_workbook(workbook_path, data_only=True, read_only=False)
    ws = workbook["PopEmpDes"]
    years = _year_columns(ws, 4)
    row_map = {
        "population_k": "População Residente",
        "labour_force_k": "População Activa - sentido lato",
        "employment_k": "Emprego Total",
        "employees_k": "Trabalhadores por conta de outrem",
        "unemployment_k": "Desemprego - sentido lato",
        "unemployment_rate": "Taxa de Desemprego - sentido lato",
    }
    resolved = {name: _row(ws, label) for name, label in row_map.items()}
    records: list[dict[str, float | int]] = []
    for year, col in sorted(years.items()):
        if not 1977 <= year <= 1995:
            continue
        record: dict[str, float | int] = {"year": year}
        for name, row in resolved.items():
            value = ws.cell(row, col).value
            if isinstance(value, (int, float)):
                record[name] = float(value)
        records.append(record)
    macro = pd.DataFrame.from_records(records)
    macro["source"] = "Banco de Portugal / INE long series"
    return macro


def _metric_rows(ws: Worksheet) -> dict[str, int]:
    """Resolve comparable public-account rows in historical sheets."""
    rows: dict[str, int] = {
        "current_revenue_m_eur": _row(ws, "Receitas correntes"),
        "social_contributions_m_eur": _row(ws, "Contribuições sociais efectivas"),
        "capital_revenue_m_eur": _row(ws, "Receitas de capital"),
        "total_revenue_m_eur": _row(ws, "Receita total"),
        "current_expenditure_m_eur": _row(ws, "Despesas correntes"),
        "compensation_m_eur": _row(ws, "Remunerações"),
        "interest_m_eur": _row(ws, "Juros"),
        "subsidies_m_eur": _row(ws, "Subsídios"),
        "capital_expenditure_m_eur": _row(ws, "Despesas de capital"),
        "gfcf_m_eur": _row(ws, "Formação bruta de capital fixo"),
        "total_expenditure_m_eur": _row(ws, "Despesa total"),
        "balance_m_eur": _row(ws, "Capacidade (+) Necessidade (-) de financiamento"),
    }
    transfer_rows = _rows_containing(ws, "Transferências correntes")
    if len(transfer_rows) >= 2:
        rows["current_transfers_received_m_eur"] = transfer_rows[0]
        rows["current_transfers_paid_m_eur"] = transfer_rows[1]
    capital_transfer_rows = _rows_containing(ws, "Transferências de capital")
    if len(capital_transfer_rows) >= 2:
        rows["capital_transfers_received_m_eur"] = capital_transfer_rows[0]
        rows["capital_transfers_paid_m_eur"] = capital_transfer_rows[1]
    return rows


def _extract_accounts(workbook_path: Path, gdp: pd.DataFrame) -> pd.DataFrame:
    workbook = load_workbook(workbook_path, data_only=True, read_only=False)
    gdp_map = dict(zip(gdp["year"].astype(int), gdp["nominal_gdp_m_eur"], strict=True))
    records: list[dict[str, float | int | str]] = []
    for sheet_name, sector in SECTOR_SHEETS.items():
        ws = workbook[sheet_name]
        years = _year_columns(ws, 4)
        rows = _metric_rows(ws)
        for year, col in sorted(years.items()):
            if not 1977 <= year <= 1995:
                continue
            record: dict[str, float | int | str] = {
                "year": year,
                "sector": sector,
                "source": "Banco de Portugal / INE long series",
                "statistical_regime": "historical_long_series",
            }
            for metric, row in rows.items():
                value = _to_m_eur(ws.cell(row, col).value)
                if value is not None:
                    record[metric] = value
            balance = record.get("balance_m_eur")
            interest = record.get("interest_m_eur")
            if isinstance(balance, float) and isinstance(interest, float):
                record["primary_balance_m_eur"] = balance + interest
            gdp_value = gdp_map.get(year)
            if gdp_value is not None:
                record["nominal_gdp_m_eur"] = float(gdp_value)
            records.append(record)
    frame = pd.DataFrame.from_records(records).sort_values(["sector", "year"])
    for column in [name for name in frame.columns if name.endswith("_m_eur") and name != "nominal_gdp_m_eur"]:
        frame[column.replace("_m_eur", "_pct_gdp")] = 100.0 * frame[column] / frame["nominal_gdp_m_eur"]
    frame["account_identity_error_m_eur"] = (
        frame["total_revenue_m_eur"] - frame["total_expenditure_m_eur"] - frame["balance_m_eur"]
    )
    return frame.reset_index(drop=True)


def _extract_transfers(workbook_path: Path) -> pd.DataFrame:
    workbook = load_workbook(workbook_path, data_only=True, read_only=False)
    records: list[dict[str, float | int | str]] = []
    for sheet_name, sector in SECTOR_SHEETS.items():
        if sheet_name == "RDAP":
            continue
        ws = workbook[sheet_name]
        years = _year_columns(ws, 4)
        rows = _rows_containing(ws, "das quais: transferências entre administrações públicas")
        if len(rows) < 4:
            continue
        for year, col in sorted(years.items()):
            if not 1977 <= year <= 1995:
                continue
            values = [_to_m_eur(ws.cell(row, col).value) for row in rows[:4]]
            if any(value is None for value in values):
                continue
            current_received, capital_received, current_paid, capital_paid = [float(v) for v in values if v is not None]
            received = current_received + capital_received
            paid = current_paid + capital_paid
            records.append(
                {
                    "year": year,
                    "sector": sector,
                    "intragov_current_received_m_eur": current_received,
                    "intragov_capital_received_m_eur": capital_received,
                    "intragov_current_paid_m_eur": current_paid,
                    "intragov_capital_paid_m_eur": capital_paid,
                    "intragov_received_m_eur": received,
                    "intragov_paid_m_eur": paid,
                    "net_intragov_transfer_m_eur": received - paid,
                    "source": "Banco de Portugal / INE long series",
                }
            )
    return pd.DataFrame.from_records(records).sort_values(["sector", "year"]).reset_index(drop=True)


def _extract_balances(workbook_path: Path, accounts: pd.DataFrame) -> pd.DataFrame:
    pivot = accounts.pivot(index="year", columns="sector", values="balance_m_eur").reset_index()
    pivot = pivot.rename(
        columns={
            "general_government": "general_government_balance_m_eur",
            "central_government": "central_government_balance_m_eur",
            "regional_local_government": "regional_local_balance_m_eur",
            "social_security_funds": "social_security_balance_m_eur",
        }
    )
    workbook = load_workbook(workbook_path, data_only=True, read_only=False)
    for sheet_name, output in ADDITIONAL_BALANCE_SHEETS.items():
        ws = workbook[sheet_name]
        years = _year_columns(ws, 4)
        row = _row(ws, "Capacidade (+) Necessidade (-) de financiamento")
        values: dict[int, float] = {}
        for year, col in years.items():
            if 1977 <= year <= 1995:
                value = _to_m_eur(ws.cell(row, col).value)
                if value is not None:
                    values[year] = value
        pivot[output] = pivot["year"].map(values)
    gdp = accounts.loc[accounts["sector"].eq("general_government"), ["year", "nominal_gdp_m_eur"]]
    pivot = pivot.merge(gdp, on="year", how="left", validate="one_to_one")
    for column in [name for name in pivot.columns if name.endswith("balance_m_eur")]:
        pivot[column.replace("_m_eur", "_pct_gdp")] = 100.0 * pivot[column] / pivot["nominal_gdp_m_eur"]
    pivot["source_primary"] = "Banco de Portugal / INE long series"
    pivot["statistical_regime"] = "historical_long_series"
    pivot["methodology_overlap_year"] = pivot["year"].eq(1995)
    return pivot.sort_values("year").reset_index(drop=True)


def extract_long_series(workbook_path: Path) -> HistoricalExtraction:
    """Extract the 1977-1995 historical bundle from the official long-series workbook."""
    gdp = _extract_gdp(workbook_path)
    accounts = _extract_accounts(workbook_path, gdp)
    balances = _extract_balances(workbook_path, accounts)
    transfers = _extract_transfers(workbook_path)
    macro = _extract_macro(workbook_path).merge(gdp, on="year", how="left", validate="one_to_one")
    return HistoricalExtraction(balances=balances, accounts=accounts, transfers=transfers, macro=macro)
