"""Parsers for Portuguese Public Finance Council (CFP) ESA 2010 workbooks."""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


@dataclass(frozen=True)
class CFPExtraction:
    """Modern headline balances, accounts and debt diagnostics."""

    general_government: pd.DataFrame
    subsectors: pd.DataFrame
    accounts: pd.DataFrame
    debt: pd.DataFrame


def _normalise(value: object) -> str:
    return " ".join(str(value or "").lower().replace("|", " ").split())


def _year_columns(ws: Worksheet, row: int = 3) -> dict[int, int]:
    result: dict[int, int] = {}
    for col in range(1, ws.max_column + 1):
        value = ws.cell(row, col).value
        if isinstance(value, (int, float)) and int(value) == value:
            year = int(value)
            if 1900 <= year <= 2100:
                result[year] = col
    if not result:
        raise ValueError(f"No annual columns found in sheet {ws.title!r}")
    return result


def _rows_containing(ws: Worksheet, label: str) -> list[int]:
    target = _normalise(label)
    rows: list[int] = []
    for row in range(1, ws.max_row + 1):
        values = [_normalise(ws.cell(row, col).value) for col in range(1, min(ws.max_column, 9) + 1)]
        if any(target in value for value in values):
            rows.append(row)
    return rows


def _row(ws: Worksheet, label: str, *, occurrence: int = 0, required: bool = True) -> int | None:
    rows = _rows_containing(ws, label)
    if occurrence < len(rows):
        return rows[occurrence]
    if required:
        raise ValueError(f"Could not find {label!r} in {ws.title!r}")
    return None


def _numeric(value: object) -> float | None:
    return float(value) if isinstance(value, (int, float)) else None


def _extract_row(ws: Worksheet, row: int) -> pd.DataFrame:
    records: list[dict[str, float | int]] = []
    for year, col in sorted(_year_columns(ws).items()):
        value = _numeric(ws.cell(row, col).value)
        if value is not None:
            records.append({"year": year, "value": value})
    return pd.DataFrame.from_records(records)


def _sheet_metrics(ws: Worksheet, *, is_general: bool) -> dict[str, int]:
    balance_label = "General government balance" if is_general else "Overall balance"
    metrics: dict[str, tuple[str, bool]] = {
        "total_revenue_m_eur": ("Total revenue", True),
        "current_revenue_m_eur": ("Current revenue", True),
        "tax_revenue_m_eur": ("Tax revenue", True),
        "social_contributions_m_eur": ("Social contributions", True),
        "sales_other_current_revenue_m_eur": ("Sales & other current rev.", True),
        "capital_revenue_m_eur": ("Capital transfers received", True),
        "total_expenditure_m_eur": ("Total expenditure", True),
        "primary_expenditure_m_eur": ("Primary expenditure", True),
        "current_primary_expenditure_m_eur": ("Current primary expend.", True),
        "intermediate_consumption_m_eur": ("Intermediate consumption", True),
        "compensation_m_eur": ("Compensation of employees", True),
        "social_transfers_m_eur": ("Social transfers", True),
        "subsidies_m_eur": ("Subsidies", True),
        "other_current_expenditure_m_eur": ("Other current expenditure", True),
        "capital_expenditure_m_eur": ("Capital expenditure", True),
        "gfcf_m_eur": ("GFCF", True),
        "interest_m_eur": ("Interest paid", True),
        "balance_m_eur": (balance_label, True),
        "primary_balance_m_eur": ("Primary balance", True),
        "nominal_gdp_m_eur": ("GDP at current market prices", True),
    }
    result: dict[str, int] = {}
    for output, (label, required) in metrics.items():
        row = _row(ws, label, required=required)
        if row is not None:
            result[output] = row
    return result


def _extract_accounts_from_sheet(ws: Worksheet, *, sector: str, is_general: bool) -> pd.DataFrame:
    rows = _sheet_metrics(ws, is_general=is_general)
    records: list[dict[str, float | int | str]] = []
    for year, col in sorted(_year_columns(ws).items()):
        record: dict[str, float | int | str] = {
            "year": year,
            "sector": sector,
            "source": "Portuguese Public Finance Council annual ESA 2010 workbook",
            "statistical_regime": "esa2010_modern",
        }
        for metric, row in rows.items():
            value = _numeric(ws.cell(row, col).value)
            if value is not None:
                record[metric] = value
        if "balance_m_eur" in record and "total_revenue_m_eur" in record and "total_expenditure_m_eur" in record:
            record["account_identity_error_m_eur"] = (
                float(record["total_revenue_m_eur"])
                - float(record["total_expenditure_m_eur"])
                - float(record["balance_m_eur"])
            )
        records.append(record)
    frame = pd.DataFrame.from_records(records)
    for column in [name for name in frame.columns if name.endswith("_m_eur") and name != "nominal_gdp_m_eur"]:
        frame[column.replace("_m_eur", "_pct_gdp")] = 100.0 * frame[column] / frame["nominal_gdp_m_eur"]
    return frame


def _extract_debt_from_sheet(ws: Worksheet, *, sector: str) -> pd.DataFrame:
    label_map = {
        "stock_flow_adjustment_m_eur": "Stock flow adjustment",
        "debt_change_m_eur": "Change in",
        "debt_m_eur": "debt  (nominal value)",
        "nominal_gdp_m_eur": "GDP at current market prices",
    }
    rows: dict[str, int] = {}
    for output, label in label_map.items():
        if output == "debt_change_m_eur":
            candidates = _rows_containing(ws, "Change in")
            candidates = [row for row in candidates if "debt" in _normalise(ws.cell(row, 3).value)]
            if candidates:
                rows[output] = candidates[0]
        elif output == "debt_m_eur":
            candidates = _rows_containing(ws, "debt")
            candidates = [
                row
                for row in candidates
                if "nominal value" in _normalise(ws.cell(row, 3).value)
                and "net of" not in _normalise(ws.cell(row, 3).value)
            ]
            if candidates:
                rows[output] = candidates[0]
        else:
            row = _row(ws, label, required=False)
            if row is not None:
                rows[output] = row
    records: list[dict[str, float | int | str]] = []
    for year, col in sorted(_year_columns(ws).items()):
        record: dict[str, float | int | str] = {"year": year, "sector": sector}
        for metric, row in rows.items():
            value = _numeric(ws.cell(row, col).value)
            if value is not None:
                record[metric] = value
        records.append(record)
    frame = pd.DataFrame.from_records(records)
    frame["source"] = "Portuguese Public Finance Council annual ESA 2010 workbook"
    return frame


def extract_cfp_annual(general_path: Path, subsector_path: Path) -> CFPExtraction:
    """Extract modern ESA 2010 balances, account components and debt diagnostics."""
    general_wb = load_workbook(general_path, data_only=True, read_only=False)
    subsector_wb = load_workbook(subsector_path, data_only=True, read_only=False)

    gg_ws = general_wb["AP S13 (M€)"]
    sector_sheets = {
        "general_government": gg_ws,
        "central_government": subsector_wb["AdC S1311 (M€)"],
        "regional_local_government": subsector_wb["ARL S1313 (M€)"],
        "social_security_funds": subsector_wb["FSS S1314 (M€)"],
    }

    account_frames = [
        _extract_accounts_from_sheet(ws, sector=sector, is_general=sector == "general_government")
        for sector, ws in sector_sheets.items()
    ]
    accounts = pd.concat(account_frames, ignore_index=True).sort_values(["sector", "year"])

    gg_accounts = accounts.loc[accounts["sector"].eq("general_government")].copy()
    general = gg_accounts[
        ["year", "balance_m_eur", "balance_pct_gdp", "nominal_gdp_m_eur"]
    ].rename(
        columns={
            "balance_m_eur": "general_government_balance_m_eur",
            "balance_pct_gdp": "general_government_balance_pct_gdp",
        }
    )
    general["source_secondary"] = "Portuguese Public Finance Council annual ESA 2010 workbook"

    sub = accounts.loc[~accounts["sector"].eq("general_government"), ["year", "sector", "balance_m_eur", "balance_pct_gdp"]]
    sub_m = sub.pivot(index="year", columns="sector", values="balance_m_eur").reset_index().rename(
        columns={
            "central_government": "central_government_balance_m_eur",
            "regional_local_government": "regional_local_balance_m_eur",
            "social_security_funds": "social_security_balance_m_eur",
        }
    )
    sub_pct = sub.pivot(index="year", columns="sector", values="balance_pct_gdp").reset_index().rename(
        columns={
            "central_government": "central_government_balance_pct_gdp",
            "regional_local_government": "regional_local_balance_pct_gdp",
            "social_security_funds": "social_security_balance_pct_gdp",
        }
    )
    subsectors = sub_m.merge(sub_pct, on="year", validate="one_to_one")
    subsectors["source_secondary"] = "Portuguese Public Finance Council annual ESA 2010 workbook"

    debt = pd.concat(
        [_extract_debt_from_sheet(ws, sector=sector) for sector, ws in sector_sheets.items()],
        ignore_index=True,
    ).sort_values(["sector", "year"])
    balance_lookup = accounts[["year", "sector", "balance_m_eur"]]
    debt = debt.merge(balance_lookup, on=["year", "sector"], how="left", validate="one_to_one")
    debt["debt_reconciliation_error_m_eur"] = (
        debt["debt_change_m_eur"] + debt["balance_m_eur"] - debt["stock_flow_adjustment_m_eur"]
    )

    return CFPExtraction(
        general_government=general.sort_values("year").reset_index(drop=True),
        subsectors=subsectors.sort_values("year").reset_index(drop=True),
        accounts=accounts.reset_index(drop=True),
        debt=debt.reset_index(drop=True),
    )


def extract_social_security_detail(workbook_path: Path) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Extract 2019-2025 system balances and selected 2024-2025 Social Security details."""
    workbook = load_workbook(workbook_path, data_only=True, read_only=False)

    chart = workbook["Gráfico 13"]
    years = [int(chart.cell(5, col).value) for col in range(3, 10)]
    system_rows = {
        "citizenship_system_balance_m_eur": 7,
        "previdential_system_balance_m_eur": 8,
        "special_regimes_balance_m_eur": 9,
    }
    system_records: list[dict[str, float | int]] = []
    for col, year in enumerate(years, start=3):
        record: dict[str, float | int] = {"year": year}
        for column, row in system_rows.items():
            value = chart.cell(row, col).value
            if not isinstance(value, (int, float)):
                raise TypeError(f"Unexpected Social Security chart value: {value!r}")
            record[column] = float(value)
        system_records.append(record)
    system_balances = pd.DataFrame.from_records(system_records)

    q1 = workbook["Quadro 1"]

    def row_by_label(ws: Worksheet, label: str) -> int:
        for row in range(1, ws.max_row + 1):
            if str(ws.cell(row, 2).value or "").strip() == label:
                return row
        raise ValueError(f"Could not find {label!r} in {ws.title!r}")

    q1_metrics = {
        "social_contributions_m_eur": "Contribuições e quotizações",
        "state_budget_transfers_m_eur": "Transferências do OE - das quais:",
        "state_budget_lbss_transfers_m_eur": "Transf. do OE para cumprimento da LBSS",
        "rsi_expenditure_m_eur": "Rendimento Social de Inserção",
        "professional_training_subsidies_m_eur": "Subsídios de Formação Profissional",
        "social_security_budget_balance_m_eur": "Saldo global (excl. FSE e FEAC)",
    }
    detail_records: list[dict[str, float | int]] = []
    for year, col in ((2024, 4), (2025, 5)):
        base_detail_record: dict[str, float | int] = {"year": year}
        for output, label in q1_metrics.items():
            value = q1.cell(row_by_label(q1, label), col).value
            if isinstance(value, (int, float)):
                base_detail_record[output] = float(value)
        detail_records.append(base_detail_record)

    q2 = workbook["Quadro 2"]
    for year, col in ((2024, 3), (2025, 6)):
        detail_record = next(item for item in detail_records if item["year"] == year)
        row_map = {
            "previdential_revenue_m_eur": 8,
            "previdential_contributions_m_eur": 9,
            "previdential_state_transfers_m_eur": 10,
            "previdential_training_employment_expenditure_m_eur": 21,
            "previdential_balance_m_eur": 24,
            "citizenship_revenue_m_eur": 28,
            "citizenship_state_lbss_transfers_m_eur": 29,
            "citizenship_expenditure_m_eur": 36,
            "citizenship_balance_m_eur": 55,
        }
        for output, row in row_map.items():
            value = q2.cell(row, col).value
            if isinstance(value, (int, float)):
                detail_record[output] = float(value)

    detail = pd.DataFrame.from_records(detail_records)
    system_balances["source"] = "CFP Social Security 2025 report underlying data"
    detail["source"] = "CFP Social Security 2025 report underlying data"
    return system_balances, detail
